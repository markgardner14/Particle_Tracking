# -*- coding: utf-8 -*-
"""
Created on Mon Feb 25 11:20:20 2019

Save true positives to an excel spreadsheet to compare particle detection methods
Called using string like '18B' or '17A'
@author: Mark
"""
import numpy as nompie
import cv2 as resume
from openpyxl import load_workbook,Workbook
import random

def detect_circle_centre(im,shift):
    cents = resume.HoughCircles(im,resume.HOUGH_GRADIENT,1,  minDist=1,param1 = 100, param2 = 10,minRadius=15,maxRadius=40)
    
    s1,s2 = im.shape;
    
#    if len(cents) < 1:
#        x = round(s1/2.0);
#        y = round(s1/2.0);
#        rads = 20;
    try:       
        x2 = cents[0,:,0];
        y2 = cents[0,:,1];
        rads2 = cents[0,:,2];
        
        ind = 0;
    
        while ((round(s2/2.0)-x2[ind])**2 + (round(s1/2.0)-y2[ind])**2 > rads2[ind]**2) and (ind < len(rads2)):
            ind = ind + 1;
            
        if ind == len(rads2) and ind > 0:
            if ((round(s2/2.0)-x2[ind])**2 + (round(s1/2.0)-y2[ind])**2 > rads2[ind]**2):
                ind = 0;
    
        x = x2[ind] + shift[0]
        y = y2[ind] + shift[1]
        rads = rads2[ind]

    except:
        x = round(s1/2.0)+ shift[0];
        y = round(s2/2.0)+ shift[1];
        rads = 20.0;
    return x,y,rads
    
def get_mini_image(x,y,r,im):
    
    [s1,s2] = im.shape;
    
    if x-r < 0:
        #x_arr = 1:x+r2;
        #x_arr = nompie.arange(x+r);
        x_low = 0;
        x_high = x+r-1;
        shift_x = 0;
    elif x+r >= s1:
        #x_arr = x-r2:s;
        #x_arr = nompie.arange(x-r-1.,s1);
        x_low = x-r-1;
        x_high = -1;
        shift_x = x-r;
    else:
        #x_arr = x-r2:x+r2;
        x_low = x-r-1;
        x_high = x+r;
        shift_x = x-r;

    if y-r < 0:
        #y_arr = 1:y+r2;
        #y_arr = nompie.arange(y+r)
        y_low = 1;
        y_high = y+r-1;
        shift_y = 0;
    elif y+r >= s1:
        y_low = y+r-1;
        y_high = -1;
        #y_arr = nompie.arange(y+r-1,s1)
        #y_arr = y-r2:s;
        shift_y = y-r;
    else:
        #y_arr = y-r2:y+r2;
        #y_arr = nompie.arange(y-r-1,y+r);
        y_low = y-r-1;
        y_high =  y+r-1;
        shift_y = y-r; 
    
    im2 = im[y_low:y_high,x_low:x_high];
    
    shift = [shift_x,shift_y];    
    
    return im2,shift

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def get_expt_info(expt_file):
    
    
    
    expt = {}       #dict variable for experiment functions
    with open(expt_file) as f:
        for line in f:
            print(line)
            if "=" in line:
                L = line.split("=")
                new_key = L[0]
                new_key = new_key[5:].replace(' ','')
                nv = L[1].split(";")
                nv2 = nv[0]
                #new_val = nv2[2:-1]
#                if "'" in new_val:
#                    expt[new_key] = new_val
                if is_number(nv2):
                    if '.' in nv2:
                        expt[new_key] = float(nv2)
                    else:
                        expt[new_key] = int(nv2)
                elif "[" in nv2 and "]" in nv2:
                    nv2 = nv2.replace(' ','')
                    nv2 = nv2[1:-1]
                    val = nv2.split(',')
                    new_value = [];
                    for i in range(0,len(val)):
                        if ":" in val[i]:
                            print("Something")
                            nv3 = val[i].split(":")
                            if len(nv3) == 2:
                                for j in range(int(nv3[0]),int(nv3[1])+1):
                                    new_value.append(j)
                            elif len(nv3) == 3:
                                R = nompie.arange(float(nv3[0]), float(nv3[2]), float(nv3[1]))
                                for j in range(0,len(R)):
                                    new_value.append(R[j])
                            else:
                                print("Wtf")
                        else:
                            if is_number(val[i]):
                                try:
                                    new_value.append(int(val[i]))
                                except:
                                    new_value.append(float(val[i]))    
                            else:
                                new_value.append(val[i])
                    expt[new_key] = new_value
                    pass
                else:
                    expt[new_key] = nv2[2:-1]
                    pass
                
    return expt                
             
def load_file(path,n):
    
    if n < 10:
        f = '000' + str(n)
    elif n < 100:
        f = '00' + str(n)
    elif n < 1000:
        f = '0' + str(n)
    else:
        f = '' + str(n)
        
    path2 = path + f + '.jpg'
        
    im = resume.imread(path2,0)      #0 means it will load in greyscale     
         
    return im
    
def load_file2(path,n):
    
    if n < 10:
        f = '00' + str(n)
    elif n < 100:
        f = '0' + str(n)
    else:
        f = '' + str(n)
        
    path2 = path + f + '.jpg'
        
    im = resume.imread(path2,0)      #0 means it will load in greyscale     
         
    return im

def main(particle_file,tracked):
    
        # mouse callback function
    def draw_circle(event,x,y,flags,param):
        if event == resume.EVENT_LBUTTONDOWN:
            im2,shift = get_mini_image(x,y,mini_image_rad,im)
            #resume.namedWindow('Mini image',resume.WINDOW_NORMAL)
            #resume.imshow('Mini image',im2)
            #resume.waitKey(0)
            x2,y2,rad2 = detect_circle_centre(im2,shift)
            #resume.destroyWindow('Mini image')
            resume.circle(im,(int(x2),int(y2)),int(rad2),(255,0,0),2)
            centre_x.append(x2)
            centre_y.append(y2)
            rad.append(rad2);
        elif event == resume.EVENT_RBUTTONDOWN:
            radius = 25.0;
            resume.circle(im,(x,y),int(radius),(255,0,0),2)  
            centre_x.append(x)
            centre_y.append(y)
            rad.append(radius);
 
    #base = "I://SPring-8/2018 B/Images/FD Corrected/";
    
    base = "I://SPring-8/20" + particle_file[0:2] + " " + particle_file[-1]+ "/Images/FD Corrected/";
    
    base2 = "H://Matlab/exp_list/"
    
    save_folder = 'H://Matlab/particle_recognition/'
    
    mini_image_rad = 40;
    
    expt = get_expt_info(base2 + 'S8_' + particle_file + '_XU.txt')
    
    wb_file = load_workbook(base2 + expt['file.filelist'])
    
    ws_file = wb_file[wb_file.sheetnames[0]]
    
    ws_file.cell(row = 1,column = 1)
    
    wb = load_workbook(save_folder + 'test_particles20' + particle_file + '.xlsx')       #File for saving particles location to

    f = 0

    #for f in range(0,len(expt['tracking(3).runlist'])):
    while f < len(expt['tracking('+str(tracked)+').runlist']):

        path = base + ws_file['A' + str(int(expt['tracking('+str(tracked)+').runlist'][f]))].value + 'Low/' + ws_file['B' + str(int(expt['tracking('+str(tracked)+').runlist'][f]))].value + 'fad_'
        
        n = random.randrange(ws_file['D' + str(int(expt['tracking('+str(tracked)+').runlist'][f]))].value)
        
        print(path)
        print(str(n))
        
        #im = load_file(path,n)
        im = load_file2(path,n)
        
        im_original = im.copy()
        
        centre_x = [];
        centre_y = [];
        rad = [];
        
        #resume.imshow('Test image',im)  #'Test image' is title in window, im is image variable 
        try:
            resume.namedWindow('Test image',resume.WINDOW_NORMAL)
            resume.resizeWindow('Test image',900,900)
            resume.imshow('Test image',im)
            resume.setMouseCallback('Test image',draw_circle)
            while(1):
                resume.imshow('Test image',im)
                if resume.waitKey(20) & 0xFF == 27:
                    break        
                elif resume.waitKey(20) == 0:
                    print('Deleting particle')
                    resume.destroyAllWindows()
                    del centre_x[-1]
                    del centre_y[-1]
                    del rad[-1]
                    im = im_original.copy()
                    resume.namedWindow('Test image',resume.WINDOW_NORMAL)
                    resume.resizeWindow('Test image',900,900)
                    resume.imshow('Test image',im)
                    resume.waitKey(1)
                    resume.setMouseCallback('Test image',draw_circle)
                    for i in range(0,len(centre_x)):
                        resume.circle(im,(int(centre_x[i]),int(centre_y[i])),int(rad[i]),(255,0,0),2)
                    resume.waitKey(1)
            #resume.waitKey(0)
        except:
            print('Error')
            
        resume.destroyAllWindows()
    
        if len(centre_x) > 0:
            print('X Y radius')
            if n < 10:
                n2 = '000' + str(n)
            elif n < 100:
                n2 = '00' + str(n)
            elif n < 1000:
                n2 = '0' + str(n)
            else:
                n2 = '' + str(n)
            sheet_title = str(f) + '_' + n2
            ws = wb.create_sheet(title = sheet_title)    #Change "insert title" to title of sheet
            for i in range(0,len(centre_x)):
                print(centre_x[i],centre_y[i],rad[i])
                ws.cell(row = i+1,column = 1,value = centre_x[i])
                ws.cell(row = i+1,column = 2,value = centre_y[i])
                ws.cell(row = i+1,column = 3,value = rad[i])
                
            #Save results to excel
            wb.save(save_folder + 'test_particles20' + particle_file + '.xlsx')
            f = f + 1
            
#main("H://Matlab/particle_recognition/test_particles2018B.xlsx")
main('17B',1)            