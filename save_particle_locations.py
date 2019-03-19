# -*- coding: utf-8 -*-
"""
Created on Thu Mar  7 09:48:14 2019

File for saving the centre pt and radius of particles. 

@author: Mark
"""

import numpy as nompie
import cv2 as resume
from openpyxl import load_workbook,Workbook
import random
import particle_stuff
import math as maths
import matplotlib.pyplot as plt
import scipy.optimize as sciopt
import scipy.stats as scist

def load_file(path,n):
    
    if n < 10:
        f = '000' + str(n)
    elif n < 100:
        f = '00' + str(n)
    elif n < 1000:
        f = '0' + str(n)
    else:
        f = '' + str(n)
        
    path2 = path + f + '.jpg'
        
    im = resume.imread(path2,0)      #0 means it will load in greyscale     
         
    return im

def main(particle_file,tracked):
    # mouse callback function
    def draw_circle(event,x,y,flags,param):
        if event == resume.EVENT_LBUTTONDOWN:
            im2,shift = get_mini_image2(x,y,mini_image_rad,mini_image_rad,im)
            #resume.namedWindow('Mini image',resume.WINDOW_NORMAL)
            #resume.imshow('Mini image',im2)
            #resume.waitKey(0)
            x2,y2,rad2 = detect_circle_centre2(im2,shift,100)
            #resume.destroyWindow('Mini image')
            resume.circle(im,(int(x2),int(y2)),int(rad2),(255,0,0),2)
            centre_x.append(x2)
            centre_y.append(y2)
            rad.append(rad2);
        elif event == resume.EVENT_RBUTTONDOWN:
            radius = 25.0;
            resume.circle(im,(x,y),int(radius),(255,0,0),2)  
            centre_x.append(x)
            centre_y.append(y)
            rad.append(radius);

    base = "I://SPring-8/20" + particle_file[0:2] + " " + particle_file[-1]+ "/Images/FD Corrected/"

    base2 = "H://Matlab/exp_list/"
    
    save_folder = 'H://Matlab/particle_recognition/'
    
    mini_image_rad = 40;
    
    expt = get_expt_info(base2 + 'S8_' + particle_file + '_XU.txt')
    
    wb_file = load_workbook(base2 + expt['file.filelist'])
    
    ws_file = wb_file[wb_file.sheetnames[0]]
    
    ws_file.cell(row = 1,column = 1)
    
    #wb = load_workbook(save_folder + 'test_particles20' + particle_file + '.xlsx')       #File for saving particles location to

    p_thresh = 300

    run_list_str = 'tracking(' + str(tracked) + ').runlist'

    f = 0

    particle_num = 0;

    for i in range(0,len(expt[run_list_str])):
        i2 = len(expt[run_list_str])-i-1
        path = base + ws_file['A' + str(int(expt[run_list_str][i2]))].value + 'Low/' + ws_file['B' + str(int(expt[run_list_str][i2]))].value + 'fad_'
        try:
            txt_file =  open(path + 'particles.txt','r')
            f = i2
            for line in txt_file:
                particle_num += 1
            txt_file.close()
            break
        except:
            pass

    while f < len(expt[run_list_str]):
        
        path = base + ws_file['A' + str(int(expt[run_list_str][f]))].value + 'Low/' + ws_file['B' + str(int(expt[run_list_str][f]))].value + 'fad_'
        
        n = random.randrange(ws_file['D' + str(int(expt[run_list_str][f]))].value)
        
        im = load_file(path,n)
        
        im_original = im.copy()
        
        centre_x = [];
        centre_y = [];
        rad = [];
        
        try:
            resume.namedWindow('Test image',resume.WINDOW_NORMAL)
            resume.resizeWindow('Test image',900,900)
            resume.imshow('Test image',im)
            resume.setMouseCallback('Test image',draw_circle)
            while(1):
                resume.imshow('Test image',im)
                if resume.waitKey(20) & 0xFF == 27:
                    break        
                elif resume.waitKey(20) == 0:
                    print('Deleting particle')
                    resume.destroyAllWindows()
                    del centre_x[-1]
                    del centre_y[-1]
                    del rad[-1]
                    im = im_original.copy()
                    resume.namedWindow('Test image',resume.WINDOW_NORMAL)
                    resume.resizeWindow('Test image',900,900)
                    resume.imshow('Test image',im)
                    resume.waitKey(1)
                    resume.setMouseCallback('Test image',draw_circle)
                    for i in range(0,len(centre_x)):
                        resume.circle(im,(int(centre_x[i]),int(centre_y[i])),int(rad[i]),(255,0,0),2)
                    resume.waitKey(1)
        except:
            print('Error')
            resume.destroyAllWindows()
            break 
       
        resume.destroyAllWindows()
        
        if len(centre_x) > 0:
            print('Frame X Y radius')
            if n < 10:
                n2 = '000' + str(n)
            elif n < 100:
                n2 = '00' + str(n)
            elif n < 1000:
                n2 = '0' + str(n)
            else:
                n2 = '' + str(n)
            #sheet_title = str(f) + '_' + n2
            #ws = wb.create_sheet(title = sheet_title)    #Change "insert title" to title of sheet
            txt_file = open(path + "particles.txt","a")
            for i in range(0,len(centre_x)):
                print(n2,centre_x[i],centre_y[i],rad[i])
                txt_file.write(str(n2) + " " + str(centre_x[i]) + " " + str(centre_y[i]) + " " + str(rad[i]) + '\n')
            txt_file.close() 
            particle_num = particle_num + len(centre_x)
            if particle_num > p_thresh:
                f = f + 1   
                particle_num = 0
        
        
main('18A',2)