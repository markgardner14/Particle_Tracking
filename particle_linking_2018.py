# -*- coding: utf-8 -*-
"""
Created on Fri Mar  1 13:17:26 2019

Read particle locations and try and link the particle together and do something magical

@author: Mark
"""

import numpy as nompie
import cv2 as resume
from openpyxl import load_workbook,Workbook
import random
import particle_stuff
import math as maths
import matplotlib.pyplot as plt
import scipy.optimize as sciopt
import scipy.stats as scist

def delete_duplicate_points(pts,dist_th):
    
    bads = [];
    
    for i in range(0,len(pts)):
#        for k in range(0,len(bads)):
#            if bads[k] == i:
#                flag = 1
#                break
#        if flag == 0:
        try:
            bads.index(i)
            continue
        except:
            dist = 0
            for j in range(0,len(pts)):
                if i == j:
                    pass
                else:
                    dist = maths.sqrt((pts[i][0] - pts[j][0])**2 + (pts[i][1] - pts[j][1])**2)
                    if dist < dist_th:
                        try:
                            bads.index(j)
                        except:
                            bads.append(j)
#        else:
#            continue
    pts_new = [];
    for i in range(0,len(pts)):
        if i in bads:
            pass
        else:
            pts_new.append(pts[i])
            
    return pts_new

def load_file(path,n):
    
    if n < 10:
        f = '000' + str(n)
    elif n < 100:
        f = '00' + str(n)
    elif n < 1000:
        f = '0' + str(n)
    else:
        f = '' + str(n)
        
    path2 = path + f + '.jpg'
        
    im = resume.imread(path2,0)      #0 means it will load in greyscale     
         
    return im


base = "I://SPring-8/2018 A/Images/FD Corrected/"
base2 = "H://Matlab/exp_list/"
save_location = "I://SPring-8/2018 A/Images/Processed/Auto_processed/"

expt = get_expt_info(base2 + "S8_18A_XU.txt")       #Get experiment information

tracked = 2;

wb = load_workbook(base2 + expt['file.filelist']) #Load experiment information from excel

ws = wb[wb.sheetnames[0]]

file_str1 = [];
file_str2 = [];
file_num = [];

i = 1;

dist_th = 40
move_th = 75

runlist = expt['tracking(' + str(tracked) + ').runlist']

for row in ws.rows:
    for j in range(0,len(runlist)):
        if i == int(runlist[j]):
            file_str1.append(row[0].value)
            file_str2.append(row[1].value)
            file_num.append(int(row[3].value))
            break
    i += 1
    
#print(file_str2)    
#print(file_num)
try:
    #Do the thing 
    for i in range(0,len(file_num)):
        file = base + file_str1[i] + 'Low/' + file_str2[i] + 'fad_'
        
        kurts = []
        
        print("Calculating blur threshold")
        
        for j in range(0,min([100,file_num[i]])):
            im = load_file(file,j+1)
            hist = resume.calcHist(im,[0],None,[256],[0,256])
            kurts.append(scist.kurtosis(hist))
            
        k = kurts[nompie.nonzero(kurts < 6 and kurts > 2.5)]
            
        blur_th = (max(k) - min(k))/4.0
            
        im = load_file(file,1)
        
        resume.namedWindow('Frame',resume.WINDOW_NORMAL)
        resume.resizeWindow('Frame',900,900)
        resume.imshow('Frame',im)    
        resume.waitKey(1)
 
        resume.namedWindow('Path',resume.WINDOW_NORMAL)
        resume.resizeWindow('Path',900,900)
        resume.imshow('Path',im)    
        resume.waitKey(1)
       
        last = 0
        
        centres_last = [];
        
        points_all = []
        
        frames = [];
        x_ROI = [];
        y_ROI = [];
        ROI_width = [];
        ROI_height = [];
        
        
        #load particle locations
        with open(save_location + file_str2[i] + 'locations.txt') as f:
            for line in f:    
                #particle_locations.append(line)
                s = line.split(' ')
                if is_number(s[0]):
                    frames.append(int(s[0]))
                    x_ROI.append(float(s[1]))
                    y_ROI.append(float(s[2]))
                    ROI_width.append(float(s[3]))
                    ROI_height.append(float(s[4]))
                    
            #print(frames)
            #print(x_ROI)
        #particle_locations.pop(0)
        #print(particle_locations)
        
        for j in range(0,file_num[i]):
            
            if j > 0:
                im = load_file(file,j+1)
                resume.imshow('Frame',im)    
                resume.waitKey(1)     
                
                im_path = im.copy()
                
                resume.imshow('Path',im_path)    
                resume.waitKey(1)                   
                
            f_temp = range(frames.index(j+1),frames.index(j+1) + frames.count(j+1))
            
            centres = [];
            
            for k in range(0,len(f_temp)):
#                x_temp.append(x_ROI[f_temp[k]])
#                y_temp.append(y_ROI[f_temp[k]])
#                w_temp.append(ROI_width[f_temp[k]])
#                h_temp.append(ROI_height[f_temp[k]])
                resume.rectangle(im,(int(x_ROI[f_temp[k]]),int(y_ROI[f_temp[k]])),(int(x_ROI[f_temp[k]] + ROI_width[f_temp[k]]),int(y_ROI[f_temp[k]] + ROI_height[f_temp[k]])),(0,255,0),2)
                [im2,shift] = get_mini_image2(x_ROI[f_temp[k]],y_ROI[f_temp[k]],ROI_height[f_temp[k]],ROI_width[f_temp[k]],im)
                [x,y,r] = detect_circle_centre2(im2,shift,150)
                centres.append([x,y])
                
            if len(centres) > 0:
                centres = delete_duplicate_points(centres,dist_th)
                
            for k in range(0,len(centres)):
                im = plot_crosses(im,centres[k],10)
               
            resume.imshow('Frame',im)    
            resume.waitKey(1)
            print("Image " + str(j+1) + " analysed")

            if len(centres_last) < 1:
                if j == 0:
                    points_all.append(centres)                  
                else:
                    pass
            else:
                #Particle Linking 
                mat = make_linking_mat(centres_last,centres,move_th)   
                row,col = sciopt.linear_sum_assignment(mat)   

                #print(row)
                #print(col)
                
                goods = []
                for k in range(0,len(points_all[last])):
                    if points_all[last][k][0] > 0:
                        goods.append(k)
                
                pts_new = nompie.zeros((len(points_all[last]),2))
                        
                for k in range(0,len(points_all[last]) + len(centres)):
                    #print(k)
                    if k >= len(centres_last):
                        if k < len(col):
                            #Particle Appeared
                            if col[k] < len(centres):
                                #pts_new = nompie.append(pts_new,centres[col[k]])
                                pts_new = nompie.concatenate((pts_new,[nompie.array([centres[col[k]][0],centres[col[k]][1]])]))
                                #pts_new.append(centres[col[k]])
                                resume.drawMarker(im_path,(int(centres[col[k]][0]),int(centres[col[k]][1])), 
                                    (255,0,0),markerType = resume.MARKER_SQUARE,markerSize = 15,thickness = 2)
                    else:
                        if len(centres) > 0:
                            if col[k] < len(centres):
                                #Particles Linked
                                #pts_new.append(centres[col[k]])
                                pts_new[goods[k],:] = centres[col[k]]
                                resume.drawMarker(im_path,(int(centres_last[k][0]),int(centres_last[k][1])), 
                                    (255,0,0),markerType = resume.MARKER_DIAMOND,markerSize = 15,thickness = 2)
                                resume.drawMarker(im_path,(int(centres[col[k]][0]),int(centres[col[k]][1])), 
                                    (255,0,0),markerType = resume.MARKER_DIAMOND,markerSize = 15,thickness = 2)
                            else:
                                #Particle disappeared 
                                pts_new[goods[k],:] = 0
                                resume.drawMarker(im_path,(int(centres_last[k][0]),int(centres_last[k][0])), 
                                    (255,0,0),markerType = resume.MARKER_CROSS,markerSize = 15,thickness = 2)                                
                        else:
                            pass
                resume.imshow('Path',im_path)    
                resume.waitKey(1)                 
                points_all.append(pts_new)    
                if j > 0:
                    for k in range(0,len(pts_new)):
                        pts_plot = [];
                        for m in range(0,j+1):
                            if k < len(points_all[m]):
                                if points_all[m][k][0] > 0:
                                    pts_plot.append(points_all[m][k])
                        
                        for m in range(0,len(pts_plot)-1):
                            resume.line(im_path,(int(pts_plot[m][0]),int(pts_plot[m][1])),(int(pts_plot[m+1][0]),int(pts_plot[m+1][1])),
                                (255,255,255),2)
                    resume.imshow('Path',im_path)    
                    resume.waitKey(1)                         
            #centres_last = points_all[j]
            centres_last = [];
            for k in range(0,len(points_all[j])):
                if points_all[j][k][0] > 0:
                    centres_last.append(points_all[j][k])
            
            last = j
    print('File Completed')
    
except:
    print('Something Fucked up')
    resume.destroyAllWindows()
    